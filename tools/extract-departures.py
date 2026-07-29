"""
Извлекает из ведомости броней (xlsx) справочник заездов и цен.

Персональные данные пассажиров НЕ извлекаются: из каждого листа берутся
только дата, транспорт, прайс и число уже занятых мест. Результат идёт в
seed/departures.json и используется для наполнения базы.

Запуск:  python3 tools/extract-departures.py <файл.xlsx>
"""
import sys, json, re, collections
import openpyxl

PLACEMENTS = ("DBL", "TWIN", "TRPL", "SNG", "SING", "SINGLE")


def norm_placement(v):
    s = str(v or "").strip().upper()
    if s in ("SING", "SINGLE", "SNG"):
        return "SNG"
    return s if s in ("DBL", "TWIN", "TRPL") else None


def parse_title(title):
    """'29.05 TAS-TZX' -> ('2026-05-29', 'TZX', False)."""
    t = title.strip()
    info = "INFO" in t.upper()
    m = re.match(r"(\d{2})\.(\d{2})\s+TAS-(TZX|BUS)", t)
    if not m:
        return None
    day, month, transport = m.group(1), m.group(2), m.group(3)
    return f"2026-{month}-{day}", transport, info


def read_price_block(ws):
    """Прайс лежит сбоку в Q:S.

    Внимание: выше прайса в тех же колонках стоит блок подсчёта номеров
    (SING/DBL/TRPL/Итого), где числа — количество номеров, а не цены.
    Читаем размещения только ниже строки «Итого», иначе счётчик номеров
    попадает в прайс (например, SING 0 превращается в цену 0).
    """
    prices, child = {}, {}
    price_block_starts = 1
    for r in range(1, 20):
        key = ws.cell(r, 17).value
        if isinstance(key, str) and key.strip().lower() == "итого":
            price_block_starts = r + 1
            break

    for r in range(1, 20):
        key = ws.cell(r, 17).value          # Q
        val = ws.cell(r, 18).value          # R
        if (r >= price_block_starts and isinstance(key, str)
                and norm_placement(key) and isinstance(val, (int, float))):
            prices[norm_placement(key)] = float(val)
        # строка с подписями детских тарифов, значения — строкой ниже
        if isinstance(key, str) and key.strip().lower().startswith("chd"):
            labels = [ws.cell(r, c).value for c in (17, 18, 19)]
            values = [ws.cell(r + 1, c).value for c in (17, 18, 19)]
            for lab, v in zip(labels, values):
                if isinstance(lab, str) and isinstance(v, (int, float)):
                    child[lab.strip()] = float(v)
    return prices, child


def count_booked(ws):
    seats = 0
    by_placement = collections.Counter()
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value and ws.cell(r, 3).value:
            seats += 1
            p = norm_placement(ws.cell(r, 7).value)
            if p:
                by_placement[p] += 1
    return seats, dict(by_placement)


def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    departures = []
    for ws in wb.worksheets:
        parsed = parse_title(ws.title)
        if not parsed:
            print(f"  пропущен лист (не разобрано название): {ws.title}", file=sys.stderr)
            continue
        date, transport, is_info = parsed
        prices, child = read_price_block(ws)
        booked, by_placement = count_booked(ws)
        departures.append({
            "code": f"{transport}{date[8:10]}{date[5:7]}",
            "date_start": date,
            "transport": transport,
            "is_info_tour": is_info,
            "prices": prices,
            "child_prices": child,
            "booked_seats": booked,
            "booked_by_placement": by_placement,
            "source_sheet": ws.title,
        })
    departures.sort(key=lambda d: (d["date_start"], d["transport"]))
    print(json.dumps(departures, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main(sys.argv[1])
